#/usr/bin/python
# -*- coding: utf-8 -*-

import logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
logger = logging.getLogger(__name__)
import os
from multiprocessing import Pool
from datetime import datetime
from pandas import DataFrame
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from prettytable import PrettyTable


class ExcelExporterChunkSizeError(Exception):
    """"Raised when the given chunk size exceeds 1048576, the maximum number of rows in an XLSX file."""
    pass


class ExcelExportProcessError(Exception):
    """Raised when an Excel export process fails."""
    def __init__(self, message, excel_export_process):
        self.message = message
        self.excel_export_process = excel_export_process


class ExcelDecoratorError(Exception):
    """Raised when an Excel decoration process fails."""
    def __init__(self, message, excel_decorator):
        self.message = message
        self.excel_decorator = excel_decorator


class ExcelExportProcessResult(object):
    """Contains the results of a successful Excel export process.

    Args:
        filepath (str): File path of the created Excel file.
        file_size (str): Human readable size of the created Excel file.
        creation_duration (str): The duration needed to create the Excel file.
        row_count (int): The count of rows exported to the Excel file.
    """
    def __init__(self, filepath, file_size, creation_duration, row_count):
        self.filepath = filepath
        self.file_size = file_size
        self.creation_duration = creation_duration
        self.row_count = row_count


class ExcelExportProcess(object):
    """Exports a chunk of records into one Excel document.

    Args:
        process_name (str): Name of the process (e.g. 'Excel exporter no. 1').
        filepath (str): File path of the Excel file to be created.
        records (list): The records to be saved in the Excel file.
        column_names (list): The column names used as header information.
        sheet_name (str): The name of the worksheet to be created in the Excel file.
        overwrite (bool): Flag to indicate whether an already existing Excel file should be overwritten or not.
    """
    def __init__(self, process_name, filepath, records, column_names, sheet_name, overwrite):
        self.process_name = process_name
        self.filepath = filepath
        self.records = records
        self.column_names = column_names
        self.sheet_name = sheet_name
        self.overwrite = overwrite

    def run(self):
        try:
            logger.info("{0} starting to export {1} rows to XLSX file {2} ...".format(self.process_name,
                                                                                  len(self.records),
                                                                                  self.filepath))
            export_start_time = datetime.now().replace(microsecond=0)

            df = DataFrame(self.records, columns=self.column_names)
            writer = ExcelWriter(self.filepath, engine='xlsxwriter', options={'encoding': 'utf-8',
                                                                              'remove_timezone': True,
                                                                              'strings_to_formulas': False})
            df.to_excel(writer, self.sheet_name)
            writer.save()

            export_end_time = datetime.now().replace(microsecond=0)
            export_duration = export_end_time - export_start_time

            file_size = self._human_readable_size(os.path.getsize(self.filepath), 2)

            logger.info("{0} exported {1} rows to XLSX file {2} ({3}) in {4}".format(self.process_name,
                                                                                     len(self.records),
                                                                                     self.filepath,
                                                                                     file_size,
                                                                                     export_duration))

            result = ExcelExportProcessResult(self.filepath, file_size, export_duration, len(self.records))
            return result
        except Exception as e:
            logger.error('{0} encountered an error: {1}'.format(self.process_name, str(e)))
            return ExcelExportProcessError(str(e), self)

    def _human_readable_size(self, size, decimal_places):
        """"Return a human readable file size."""
        for unit in ['','KB','MB','GB','TB']:
            if size < 1024.0:
                break
            size /= 1024.0
        return f"{size:.{decimal_places}f}{unit}"


class ExcelExporterResult(object):
    """Contains the results of an Excel export."""
    def __init__(self, excel_export_process_results, excel_export_process_errors):
        self.excel_export_process_results = excel_export_process_results
        self.excel_export_process_errors = excel_export_process_errors

    def has_errros(self):
        """Indicates whether the Excel export encountered errors."""
        return len(self.excel_export_process_errors) > 0


class ExcelDcorationManagerResult(object):
    """Contains the results of an Excel export."""
    def __init__(self, excel_decoration_process_results, excel_decoration_process_errors):
        self.excel_decoration_process_results = excel_decoration_process_results
        self.excel_decoration_process_errors = excel_decoration_process_errors

    def has_errros(self):
        """Indicates whether the Excel export encountered errors."""
        return len(self.excel_decoration_process_errors) > 0


class ExcelExporter(object):
    """"Exports records fetched from a PostgreSQL database into one or several Excel documents.

    Args:
        filepath (str): File path of the Excel file to be created.
        query_result (QueryResult): The query result to be exported to one or several Excel file(s).
        chunk_size (int): The maximum count of rows exported to a single Excel file.
        sheet_name (str): The name of the worksheet to be created in the Excel file(s).
        overwrite (bool): Flag to indicate whether an already existing Excel file should be overwritten or not.
        parallel_processes (int): The maximum count of parallel Excel export processes to be started, defaults to 2.
    """
    def __init__(self, filepath, query_result, chunk_size, sheet_name, overwrite=False, parallel_processes=2):
        self.filepath = filepath
        self.query_result = query_result
        self.chunk_size = chunk_size
        self.sheet_name = sheet_name
        self.overwrite = overwrite
        self.parallel_processes = parallel_processes
        self.filepath_part, self.filepath_extension = os.path.splitext(self.filepath)

        if self.chunk_size > 1048576:
            raise ExcelExporterChunkSizeError("The chunk size must not exceed 1048576, the maximum number of rows in an XLSX file.")

    def export(self):
        total_file_count = self._calculate_total_file_count()
        pool = Pool(processes=self.parallel_processes)
        process_results = []

        logger.info('+{0}+'.format(60 * '-'))
        logger.info('Excel exporter initializing a pool with {0} parallel processes for creating {1} XLSX file(s)...'.format(self.parallel_processes,
                                                                                                                             total_file_count))

        total_export_start_time = datetime.now().replace(microsecond=0)
        file_counter = 0
        for chunk_records in self._chunker(self.query_result.records, self.chunk_size):
            file_counter += 1

            if total_file_count == 1:
                xlsx_filepath = "{0}{1}".format(self.filepath_part, self.filepath_extension)
            else:
                xlsx_filepath = "{0}_{1}{2}".format(self.filepath_part, file_counter, self.filepath_extension)

            if not self.overwrite:
                if os.path.isfile(xlsx_filepath):
                    logger.info("Excel file at {0} already exists, will not overwrite it".format(xlsx_filepath))
                    continue

            excel_export_process = ExcelExportProcess('Excel export process no. {0}'.format(file_counter), xlsx_filepath,
                                                      chunk_records, self.query_result.column_names, self.sheet_name, False)

            process_result = pool.apply_async(excel_export_process.run)
            process_results.append(process_result)

        pool.close()
        pool.join()

        total_export_end_time = datetime.now().replace(microsecond=0)
        total_export_duration = total_export_end_time - total_export_start_time

        logger.info('Excel exporter finished, gathering results...')

        excel_export_process_results = []
        excel_export_process_errors = []

        for process_result in process_results:
            if isinstance(process_result.get(), ExcelExportProcessError):
                excel_export_process_error = process_result.get()
                excel_export_process_errors.append(excel_export_process_error)
            else:
                excel_export_process_result = process_result.get()
                excel_export_process_results.append(excel_export_process_result)

        excel_export_result = ExcelExporterResult(excel_export_process_results, excel_export_process_errors)

        pt = PrettyTable()
        pt.field_names = ['Statistic label', 'Statistic content']
        pt.add_row(['Excel export errors', len(excel_export_process_errors)])
        pt.add_row(['Successsfully exported files', len(excel_export_process_results)])
        pt.add_row(['Total export time', total_export_duration])
        logger.info('{0}{1}'.format('Statistics:\n', pt))

        return excel_export_result

    def _chunker(self, seq, size):
        """Chunk a sequence by the given size."""
        return (seq[pos:pos + size] for pos in range(0, len(seq), size))

    def _calculate_total_file_count(self):
        """Calculates the total count of Excel files to be created."""
        mod = self.query_result.record_count % self.chunk_size
        if mod > 0:
            total_file_count = int(self.query_result.record_count / self.chunk_size) + 1
        else:
            total_file_count = int(self.query_result.record_count / self.chunk_size)
        return total_file_count


class ExcelDecorationElement(object):
    """Stores the label and content for a single Excel decoration element.

    Args:
        label (str): The label of the decoration element, e.g. 'Created by'.
        content (str): The content of the decoration element, e.g. 'Darth Vader'.
    """
    def __init__(self, label, content):
        self.label = label
        self.content = content


class ExcelDecoration(object):
    """Stores all elements for an Excel decoration.

    Args:
        sheet_name (str): The name of the worksheet to be created.
        title (str): The title of the decoration, e.g. 'Details' or 'Info'.
    """
    def __init__(self, sheet_name, title):
        self.sheet_name = sheet_name
        self.title = title
        self.elements = []

    def add_element(self, element):
        """Adds an excel decoration element."""
        self.elements.append(element)


class ExcelDecorator(object):
    """Decorates a given Excel file with one or more decorations.

    Args:
        process_name (str): Name of the excel decoration process, e.g. 'Excel decoration process no. 1'.
        filepath (str): File path of the Excel file to be decorated.
        decorations (list): The decorations to be applied to the Excel file.
    """
    def __init__(self, process_name, filepath, decorations):
        self.process_name = process_name
        self.filepath = filepath
        self.decorations = decorations

    def add_decoration(self, decoration):
        self.decorations.append(decoration)

    def _replace_placeholders(self, content):
        if content == 'CURRENT_DATETIME':
            return datetime.now()
        else:
            return content

    def decorate(self):
        try:
            title_font = Font(name='Calibri',
                              size=22,
                              bold=False,
                              italic=False,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            key_font = Font(name='Calibri',
                            size=11,
                            bold=True,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000')

            value_font = Font(name='Calibri',
                              size=11,
                              bold=False,
                              italic=False,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            decoration_start_time = datetime.now().replace(microsecond=0)

            logger.info("{0} writing additional informations to Excel file at {1}...".format(self.process_name, self.filepath))
            wb = load_workbook(filename=self.filepath)

            for decoration in self.decorations:
                logger.info("{0} adding work sheet named '{1}' with title '{2}'".format(self.process_name, decoration.sheet_name, decoration.title))
                ws = wb.create_sheet(title=decoration.sheet_name)
                ws['B3'] = decoration.title
                ws['B3'].font = title_font

                cell_index = 6
                for element in decoration.elements:
                    key_cell = "B{0}".format(cell_index)
                    value_cell = "C{0}".format(cell_index)

                    ws[key_cell] = element.label
                    ws[key_cell].font = key_font
                    ws[key_cell].alignment = Alignment(horizontal='right')
                    ws[value_cell] = self._replace_placeholders(element.content)
                    ws[value_cell].font = value_font
                    ws[value_cell].alignment = Alignment(horizontal='left')

                    cell_index += 1

            logger.info("{0} saving Excel file at {1}...".format(self.process_name, self.filepath))
            wb.save(self.filepath)

            decoration_end_time = datetime.now().replace(microsecond=0)
            decoration_duration = decoration_end_time - decoration_start_time
            logger.info("{0} successfully decorated Excel file at {1} in {2}".format(self.process_name, self.filepath, decoration_duration))

            return self.filepath
        except Exception as e:
            logger.error('{0} encountered an error: {1}'.format(self.process_name, str(e)))
            return ExcelDecoratorError(str(e), self)


class ExcelDecorationManager(object):
    """Applies the given decorations to the given Excel files.

    Args:
        filepaths (list): A list of Excel file paths to be decorated.
        decorations (list): A list of Excel decorations to be applied to each Excel file.
        parallel_processes (int): The maximum count of parallel Excel decoration processes to be started, defaults to 2.
    """
    def __init__(self, filepaths, decorations, parallel_processes=2):
        self.filepaths = filepaths
        self.decorations = decorations
        self.parallel_processes = parallel_processes

    def decorate(self):
        logger.info('+{0}+'.format(60 * '-'))
        logger.info('Excel decoration manager initializing a pool with {0} parallel processes for decorating {1} XLSX file(s)...'.format(self.parallel_processes,
                                                                                                                             len(self.filepaths)))

        pool = Pool(processes=self.parallel_processes)
        process_results = []

        total_decoration_start_time = datetime.now().replace(microsecond=0)
        file_counter = 0
        for filepath in self.filepaths:
            file_counter += 1
            process_name = 'Excel decoration process {0}'.format(file_counter)
            excel_decorator = ExcelDecorator(process_name, filepath, self.decorations)
            process_result = pool.apply_async(excel_decorator.decorate)
            process_results.append(process_result)

        pool.close()
        pool.join()

        total_decoration_end_time = datetime.now().replace(microsecond=0)
        total_decoration_duration = total_decoration_end_time - total_decoration_start_time

        logger.info('Excel decoration finished, gathering results...')

        excel_decoration_process_results = []
        excel_decoration_process_errors = []

        for process_result in process_results:
            if isinstance(process_result.get(), ExcelDecoratorError):
                excel_decoration_process_error = process_result.get()
                excel_decoration_process_errors.append(excel_decoration_process_error)
            else:
                excel_decoration_process_result = process_result.get()
                excel_decoration_process_results.append(excel_decoration_process_result)

        pt = PrettyTable()
        pt.field_names = ['Statistic label', 'Statistic content']
        pt.add_row(['Excel decoration errors', len(excel_decoration_process_errors)])
        pt.add_row(['Successsfully decorated files', len(excel_decoration_process_results)])
        pt.add_row(['Total decoration time', total_decoration_duration])
        logger.info('{0}{1}'.format('Statistics:\n', pt))

        return ExcelDcorationManagerResult(excel_decoration_process_results, excel_decoration_process_errors)


class SQLFileWriter(object):
    """Write an SQL query to a text file.

    Args:
        filepath (str): File path of the SQL file to be created.
        query (str): SQL query to be written to the file.
    """
    def __init__(self, filepath, sql_statement):
        self.filepath = filepath
        self.sql_statement = sql_statement

    def write(self):
        """Write SQL query to a file."""
        with open(self.filepath, mode='w', encoding='utf-8') as a_file:
            a_file.write(self.sql_statement)